"""
Glossary Service – manages glossaries, terms, XLSX import/export, and
usage analytics.

Tables (Supabase):
  glossaries         – glossary sets (name, source_locale, target_locales, ...)
  glossary_terms     – terms inside a glossary (source_term, pos, description, ...)
  glossary_term_translations – per-language translations for each term
  glossary_usage_log – tracks which terms were applied during translation runs

Usage analytics are computed from:
  1) article content scans (how many articles contain the term)
  2) glossary_usage_log  (how many translation runs applied the term)
"""

import re
import uuid
import time
import requests
from io import BytesIO
from datetime import datetime, timezone
from typing import Dict, List, Optional, Tuple

from config import SUPABASE_URL, SUPABASE_SERVICE_KEY, TARGET_LANGUAGES, BASE_LANGUAGE

REST_BASE = f"{SUPABASE_URL.rstrip('/')}/rest/v1" if SUPABASE_URL else ""

# Table names
GLOSSARIES_TABLE = "glossaries"
TERMS_TABLE = "glossary_terms"
TERM_TRANSLATIONS_TABLE = "glossary_term_translations"
USAGE_LOG_TABLE = "glossary_usage_log"
PULL_TABLE = "pull_registry"

# Setup SQL for all glossary tables
SETUP_SQL = """
-- Glossary sets
CREATE TABLE IF NOT EXISTS glossaries (
    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    name TEXT NOT NULL,
    source_locale TEXT NOT NULL DEFAULT 'en',
    target_locales JSONB NOT NULL DEFAULT '[]'::jsonb,
    created_by TEXT NOT NULL DEFAULT 'system',
    is_active BOOLEAN NOT NULL DEFAULT true,
    created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
    updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
);

-- Glossary terms
CREATE TABLE IF NOT EXISTS glossary_terms (
    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    glossary_id UUID NOT NULL REFERENCES glossaries(id) ON DELETE CASCADE,
    source_term TEXT NOT NULL,
    part_of_speech TEXT DEFAULT '',
    description TEXT DEFAULT '',
    image_url TEXT DEFAULT '',
    is_active BOOLEAN NOT NULL DEFAULT true,
    created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
    updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
    UNIQUE(glossary_id, source_term)
);

-- Per-language translations for each term
CREATE TABLE IF NOT EXISTS glossary_term_translations (
    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    term_id UUID NOT NULL REFERENCES glossary_terms(id) ON DELETE CASCADE,
    locale TEXT NOT NULL,
    translated_term TEXT NOT NULL DEFAULT '',
    created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
    updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
    UNIQUE(term_id, locale)
);

-- Usage tracking log
CREATE TABLE IF NOT EXISTS glossary_usage_log (
    id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    term_id UUID NOT NULL REFERENCES glossary_terms(id) ON DELETE CASCADE,
    glossary_id UUID NOT NULL REFERENCES glossaries(id) ON DELETE CASCADE,
    article_intercom_id TEXT NOT NULL,
    locale TEXT NOT NULL,
    matched_count INT NOT NULL DEFAULT 0,
    applied_at TIMESTAMPTZ NOT NULL DEFAULT now()
);

-- Indexes
CREATE INDEX IF NOT EXISTS idx_glossary_terms_glossary ON glossary_terms(glossary_id);
CREATE INDEX IF NOT EXISTS idx_glossary_term_trans_term ON glossary_term_translations(term_id);
CREATE INDEX IF NOT EXISTS idx_glossary_usage_term ON glossary_usage_log(term_id);
CREATE INDEX IF NOT EXISTS idx_glossary_usage_article ON glossary_usage_log(article_intercom_id);
"""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _headers(prefer: str = "") -> Dict[str, str]:
    if not SUPABASE_SERVICE_KEY:
        raise ValueError("SUPABASE_SERVICE_KEY must be set")
    h = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
    }
    if prefer:
        h["Prefer"] = prefer
    return h


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def tables_exist() -> bool:
    """Check if the glossaries table exists."""
    if not REST_BASE:
        return False
    try:
        h = _headers()
        h.pop("Prefer", None)
        resp = requests.get(
            f"{REST_BASE}/{GLOSSARIES_TABLE}",
            headers=h,
            params={"select": "id", "limit": "1"},
            timeout=10,
        )
        return resp.status_code != 404
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Glossary CRUD
# ---------------------------------------------------------------------------

def list_glossaries(
    search: str = "",
    status_filter: str = "ALL",
    sort_by: str = "name_asc",
    page: int = 1,
    page_size: int = 25,
) -> Dict:
    """List glossaries with term counts, filtering, search, sort, and pagination."""
    if not REST_BASE:
        return {"glossaries": [], "total": 0, "page": page, "page_size": page_size}
    h = _headers()
    h.pop("Prefer", None)

    # Build query params
    params: Dict = {"select": "*"}
    
    # Status filter
    if status_filter == "ACTIVE":
        params["is_active"] = "eq.true"
    elif status_filter == "INACTIVE":
        params["is_active"] = "eq.false"
    
    # Search (name)
    if search:
        params["name"] = f"ilike.*{search}*"
    
    # Sort
    sort_map = {
        "name_asc": "name.asc",
        "name_desc": "name.desc",
        "created_desc": "created_at.desc",
        "created_asc": "created_at.asc",
    }
    params["order"] = sort_map.get(sort_by, "name.asc")
    
    # Get all matching glossaries (we'll paginate after getting term counts)
    try:
        resp = requests.get(
            f"{REST_BASE}/{GLOSSARIES_TABLE}",
            headers=h,
            params=params,
            timeout=15,
        )
        if not resp.ok:
            return {"glossaries": [], "total": 0, "page": page, "page_size": page_size}
        all_glossaries = resp.json()
        if not isinstance(all_glossaries, list):
            all_glossaries = []
    except Exception:
        return {"glossaries": [], "total": 0, "page": page, "page_size": page_size}

    # Get term counts per glossary
    for g in all_glossaries:
        gid = g.get("id", "")
        try:
            t_resp = requests.get(
                f"{REST_BASE}/{TERMS_TABLE}",
                headers=h,
                params={
                    "select": "id",
                    "glossary_id": f"eq.{gid}",
                    "is_active": "eq.true",
                },
                timeout=10,
            )
            if t_resp.ok:
                terms = t_resp.json()
                g["term_count"] = len(terms) if isinstance(terms, list) else 0
            else:
                g["term_count"] = 0
        except Exception:
            g["term_count"] = 0

    # Sort by term count if needed
    if sort_by == "terms_desc":
        all_glossaries.sort(key=lambda x: x.get("term_count", 0), reverse=True)

    # Paginate
    total = len(all_glossaries)
    start = (page - 1) * page_size
    end = start + page_size
    page_glossaries = all_glossaries[start:end]

    return {
        "glossaries": page_glossaries,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


def create_glossary(name: str, source_locale: str, target_locales: List[str], created_by: str = "system") -> Dict:
    """Create a new glossary."""
    if not REST_BASE:
        raise ValueError("SUPABASE_URL must be set")
    row = {
        "name": name.strip(),
        "source_locale": source_locale or BASE_LANGUAGE,
        "target_locales": target_locales,
        "created_by": created_by,
        "is_active": True,
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    h = _headers("return=representation")
    resp = requests.post(f"{REST_BASE}/{GLOSSARIES_TABLE}", json=row, headers=h, timeout=15)
    if not resp.ok:
        raise RuntimeError(f"Failed to create glossary: {resp.status_code} {resp.text[:300]}")
    data = resp.json()
    return data[0] if isinstance(data, list) and data else data


def update_glossary(glossary_id: str, updates: Dict) -> Dict:
    """Update glossary settings (name, target_locales, etc.).
    Translations for deselected languages are preserved (hidden, not deleted).
    Re-selecting a language restores its column with existing data.
    """
    if not REST_BASE:
        raise ValueError("SUPABASE_URL must be set")

    # When target_locales are changed, we do NOT delete translations for removed languages.
    # Instead, the translations are preserved in the database (archived in place).
    # The frontend hides columns for deselected languages.
    # If the user re-selects a language, its column reappears with existing data intact.

    allowed = {"name", "source_locale", "target_locales", "is_active"}
    patch = {k: v for k, v in updates.items() if k in allowed}
    patch["updated_at"] = _now_iso()
    h = _headers("return=representation")
    resp = requests.patch(
        f"{REST_BASE}/{GLOSSARIES_TABLE}?id=eq.{glossary_id}",
        json=patch,
        headers=h,
        timeout=15,
    )
    if not resp.ok:
        raise RuntimeError(f"Failed to update glossary: {resp.status_code} {resp.text[:300]}")
    data = resp.json()
    return data[0] if isinstance(data, list) and data else data


def _archive_translations_for_locales(glossary_id: str, removed_locales: List[str]):
    """Delete term translations for removed target languages from a glossary."""
    if not REST_BASE or not removed_locales:
        return
    h = _headers()
    h.pop("Prefer", None)
    try:
        # Get all active term IDs for this glossary
        resp = requests.get(
            f"{REST_BASE}/{TERMS_TABLE}",
            headers=h,
            params={
                "select": "id",
                "glossary_id": f"eq.{glossary_id}",
                "is_active": "eq.true",
                "limit": "50000",
            },
            timeout=15,
        )
        if not resp.ok:
            return
        terms = resp.json()
        if not isinstance(terms, list) or not terms:
            return
        term_ids = [t["id"] for t in terms]

        # Delete translations for removed locales
        for locale in removed_locales:
            ids_str = ",".join(term_ids)
            requests.delete(
                f"{REST_BASE}/{TERM_TRANSLATIONS_TABLE}?term_id=in.({ids_str})&locale=eq.{locale}",
                headers=_headers(),
                timeout=30,
            )
    except Exception:
        pass


def delete_glossary(glossary_id: str) -> bool:
    """Hard-delete a glossary and all its terms + translations."""
    if not REST_BASE:
        return False
    h = _headers()
    h.pop("Prefer", None)
    try:
        # 1) Get all term IDs for this glossary
        resp = requests.get(
            f"{REST_BASE}/{TERMS_TABLE}",
            headers=h,
            params={"select": "id", "glossary_id": f"eq.{glossary_id}", "limit": "50000"},
            timeout=15,
        )
        term_ids = [t["id"] for t in (resp.json() if resp.ok else [])]

        # 2) Delete term translations for those terms
        if term_ids:
            ids_str = ",".join(term_ids)
            requests.delete(
                f"{REST_BASE}/{TERM_TRANSLATIONS_TABLE}?term_id=in.({ids_str})",
                headers=h,
                timeout=15,
            )

        # 3) Delete terms
        requests.delete(
            f"{REST_BASE}/{TERMS_TABLE}?glossary_id=eq.{glossary_id}",
            headers=h,
            timeout=15,
        )

        # 4) Delete the glossary itself
        resp = requests.delete(
            f"{REST_BASE}/{GLOSSARIES_TABLE}?id=eq.{glossary_id}",
            headers=h,
            timeout=15,
        )
        return resp.ok
    except Exception:
        return False


def get_glossary(glossary_id: str) -> Optional[Dict]:
    """Get a single glossary by ID."""
    if not REST_BASE:
        return None
    h = _headers()
    h.pop("Prefer", None)
    try:
        resp = requests.get(
            f"{REST_BASE}/{GLOSSARIES_TABLE}",
            headers=h,
            params={"select": "*", "id": f"eq.{glossary_id}"},
            timeout=10,
        )
        if resp.ok:
            data = resp.json()
            if isinstance(data, list) and len(data) > 0:
                return data[0]
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Term CRUD
# ---------------------------------------------------------------------------

def list_terms(
    glossary_id: str,
    search: str = "",
    page: int = 1,
    page_size: int = 25,
) -> Dict:
    """List terms in a glossary with their translations, paginated."""
    if not REST_BASE:
        return {"terms": [], "total": 0, "page": 1, "page_size": page_size}

    h = _headers()
    h.pop("Prefer", None)

    # Fetch all active terms for this glossary
    params: Dict = {
        "select": "*",
        "glossary_id": f"eq.{glossary_id}",
        "is_active": "eq.true",
        "order": "source_term.asc",
        "limit": "5000",
    }
    try:
        resp = requests.get(f"{REST_BASE}/{TERMS_TABLE}", headers=h, params=params, timeout=15)
        if not resp.ok:
            return {"terms": [], "total": 0, "page": 1, "page_size": page_size}
        all_terms = resp.json()
        if not isinstance(all_terms, list):
            all_terms = []
    except Exception:
        return {"terms": [], "total": 0, "page": 1, "page_size": page_size}

    # Fetch all translations for these terms
    term_ids = [t["id"] for t in all_terms if t.get("id")]
    trans_map: Dict[str, List[Dict]] = {}
    if term_ids:
        try:
            # Batch fetch (might be large)
            t_resp = requests.get(
                f"{REST_BASE}/{TERM_TRANSLATIONS_TABLE}",
                headers=h,
                params={
                    "select": "id,term_id,locale,translated_term",
                    "limit": "50000",
                },
                timeout=20,
            )
            if t_resp.ok:
                all_trans = t_resp.json()
                if isinstance(all_trans, list):
                    for tr in all_trans:
                        tid = tr.get("term_id", "")
                        if tid:
                            trans_map.setdefault(tid, []).append(tr)
        except Exception:
            pass

    # Enrich terms with translations dict
    for term in all_terms:
        tid = term.get("id", "")
        translations = trans_map.get(tid, [])
        term["translations"] = {t["locale"]: t["translated_term"] for t in translations if t.get("locale")}

    # Search filter
    if search:
        search_lower = search.lower()
        filtered = []
        for term in all_terms:
            if search_lower in (term.get("source_term") or "").lower():
                filtered.append(term)
                continue
            # Also search in translations
            for loc, trans in term.get("translations", {}).items():
                if search_lower in (trans or "").lower():
                    filtered.append(term)
                    break
        all_terms = filtered

    total = len(all_terms)
    start = (page - 1) * page_size
    end = start + page_size
    page_terms = all_terms[start:end]

    return {
        "terms": page_terms,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


def create_term(
    glossary_id: str,
    source_term: str,
    translations: Dict[str, str],
    part_of_speech: str = "",
    description: str = "",
    image_url: str = "",
) -> Dict:
    """Create a new term with translations."""
    if not REST_BASE:
        raise ValueError("SUPABASE_URL must be set")

    row = {
        "glossary_id": glossary_id,
        "source_term": source_term.strip(),
        "part_of_speech": part_of_speech,
        "description": description,
        "image_url": image_url,
        "is_active": True,
        "created_at": _now_iso(),
        "updated_at": _now_iso(),
    }
    h = _headers("return=representation")
    resp = requests.post(f"{REST_BASE}/{TERMS_TABLE}", json=row, headers=h, timeout=15)
    if not resp.ok:
        raise RuntimeError(f"Failed to create term: {resp.status_code} {resp.text[:300]}")
    data = resp.json()
    term = data[0] if isinstance(data, list) and data else data
    term_id = term.get("id", "")

    # Save translations
    if term_id and translations:
        _save_term_translations(term_id, translations)

    term["translations"] = translations
    return term


def update_term(
    term_id: str,
    updates: Dict,
    translations: Optional[Dict[str, str]] = None,
) -> Dict:
    """Update a term and its translations."""
    if not REST_BASE:
        raise ValueError("SUPABASE_URL must be set")

    allowed = {"source_term", "part_of_speech", "description", "image_url"}
    patch = {k: v for k, v in updates.items() if k in allowed}
    patch["updated_at"] = _now_iso()

    h = _headers("return=representation")
    resp = requests.patch(
        f"{REST_BASE}/{TERMS_TABLE}?id=eq.{term_id}",
        json=patch,
        headers=h,
        timeout=15,
    )
    if not resp.ok:
        raise RuntimeError(f"Failed to update term: {resp.status_code} {resp.text[:300]}")
    data = resp.json()
    term = data[0] if isinstance(data, list) and data else data

    if translations is not None:
        _save_term_translations(term_id, translations)

    term["translations"] = translations or {}
    return term


def delete_terms(term_ids: List[str]) -> int:
    """Soft-delete terms (set is_active=false)."""
    if not REST_BASE or not term_ids:
        return 0
    h = _headers()
    count = 0
    for tid in term_ids:
        try:
            resp = requests.patch(
                f"{REST_BASE}/{TERMS_TABLE}?id=eq.{tid}",
                json={"is_active": False, "updated_at": _now_iso()},
                headers=h,
                timeout=10,
            )
            if resp.ok:
                count += 1
        except Exception:
            pass
    return count


def _save_term_translations(term_id: str, translations: Dict[str, str]):
    """Upsert translations for a term."""
    h = _headers("resolution=merge-duplicates,return=minimal")
    for locale, translated in translations.items():
        if not locale:
            continue
        row = {
            "term_id": term_id,
            "locale": locale,
            "translated_term": translated or "",
            "updated_at": _now_iso(),
        }
        try:
            resp = requests.post(
                f"{REST_BASE}/{TERM_TRANSLATIONS_TABLE}",
                json=row,
                headers=h,
                timeout=10,
            )
            if not resp.ok and resp.status_code == 409:
                # Update existing
                requests.patch(
                    f"{REST_BASE}/{TERM_TRANSLATIONS_TABLE}?term_id=eq.{term_id}&locale=eq.{locale}",
                    json={"translated_term": translated or "", "updated_at": _now_iso()},
                    headers=_headers(),
                    timeout=10,
                )
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Usage Analytics
# ---------------------------------------------------------------------------

def compute_term_usage(glossary_id: str) -> Dict[str, Dict]:
    """
    Compute usage analytics for all terms in a glossary.
    Returns: { term_id: { article_count, translation_count } }
    """
    if not REST_BASE:
        return {}

    h = _headers()
    h.pop("Prefer", None)

    # Get terms
    try:
        resp = requests.get(
            f"{REST_BASE}/{TERMS_TABLE}",
            headers=h,
            params={
                "select": "id,source_term",
                "glossary_id": f"eq.{glossary_id}",
                "is_active": "eq.true",
                "limit": "5000",
            },
            timeout=15,
        )
        if not resp.ok:
            return {}
        terms = resp.json()
        if not isinstance(terms, list):
            return {}
    except Exception:
        return {}

    if not terms:
        return {}

    # Get pulled article content for scanning
    articles_content = _get_pulled_articles_content()

    # Get usage log entries
    usage_log: Dict[str, int] = {}
    try:
        log_resp = requests.get(
            f"{REST_BASE}/{USAGE_LOG_TABLE}",
            headers=h,
            params={
                "select": "term_id",
                "glossary_id": f"eq.{glossary_id}",
                "limit": "50000",
            },
            timeout=15,
        )
        if log_resp.ok:
            logs = log_resp.json()
            if isinstance(logs, list):
                for entry in logs:
                    tid = entry.get("term_id", "")
                    if tid:
                        usage_log[tid] = usage_log.get(tid, 0) + 1
    except Exception:
        pass

    # Compute article_count (how many articles contain this term)
    result: Dict[str, Dict] = {}
    for term in terms:
        tid = term.get("id", "")
        source = (term.get("source_term") or "").strip()
        if not source:
            continue

        article_count = 0
        pattern = re.compile(re.escape(source), re.IGNORECASE)
        for art in articles_content:
            text = (art.get("title") or "") + " " + (art.get("body_html") or "")
            if pattern.search(text):
                article_count += 1

        result[tid] = {
            "article_count": article_count,
            "translation_count": usage_log.get(tid, 0),
        }

    return result


def _get_pulled_articles_content() -> List[Dict]:
    """Get all pulled articles with title + body_html for term scanning."""
    if not REST_BASE:
        return []
    h = _headers()
    h.pop("Prefer", None)
    try:
        resp = requests.get(
            f"{REST_BASE}/{PULL_TABLE}",
            headers=h,
            params={
                "select": "intercom_id,title,body_html",
                "pulled_at": "not.is.null",
                "limit": "5000",
            },
            timeout=20,
        )
        if resp.ok:
            data = resp.json()
            return data if isinstance(data, list) else []
    except Exception:
        pass
    return []


def log_glossary_usage(
    term_id: str,
    glossary_id: str,
    article_intercom_id: str,
    locale: str,
    matched_count: int = 1,
):
    """Log that a glossary term was applied during translation."""
    if not REST_BASE:
        return
    h = _headers("return=minimal")
    row = {
        "term_id": term_id,
        "glossary_id": glossary_id,
        "article_intercom_id": article_intercom_id,
        "locale": locale,
        "matched_count": matched_count,
        "applied_at": _now_iso(),
    }
    try:
        requests.post(f"{REST_BASE}/{USAGE_LOG_TABLE}", json=row, headers=h, timeout=10)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Glossary <-> Translation Integration
# ---------------------------------------------------------------------------

def get_active_glossary_terms(glossary_id: Optional[str] = None) -> List[Dict]:
    """
    Get all active glossary terms with their translations.
    If glossary_id is None, fetches terms from ALL active glossaries.
    If glossary_id is provided, fetches terms only from that glossary.
    Returns: [{ id, glossary_id, source_term, translations: {locale: translated_term} }]
    """
    if not REST_BASE:
        return []
    h = _headers()
    h.pop("Prefer", None)

    # Determine which glossary IDs to pull terms from
    glossary_ids: List[str] = []
    if glossary_id:
        glossary_ids = [glossary_id]
    else:
        # Fetch ALL active glossary IDs
        try:
            resp = requests.get(
                f"{REST_BASE}/{GLOSSARIES_TABLE}",
                headers=h,
                params={
                    "select": "id",
                    "is_active": "eq.true",
                    "limit": "500",
                },
                timeout=10,
            )
            if resp.ok:
                data = resp.json()
                if isinstance(data, list):
                    glossary_ids = [g["id"] for g in data]
        except Exception:
            pass

    if not glossary_ids:
        return []

    # Get terms from all active glossaries
    all_terms: List[Dict] = []
    try:
        ids_filter = ",".join(glossary_ids)
        resp = requests.get(
            f"{REST_BASE}/{TERMS_TABLE}",
            headers=h,
            params={
                "select": "id,glossary_id,source_term",
                "glossary_id": f"in.({ids_filter})",
                "is_active": "eq.true",
                "limit": "10000",
            },
            timeout=15,
        )
        if not resp.ok:
            return []
        terms = resp.json()
        if isinstance(terms, list):
            all_terms = terms
    except Exception:
        return []

    if not all_terms:
        return []

    # Get translations for all terms
    term_ids = [t["id"] for t in all_terms]
    terms = all_terms
    trans_map: Dict[str, Dict[str, str]] = {}
    if term_ids:
        try:
            t_resp = requests.get(
                f"{REST_BASE}/{TERM_TRANSLATIONS_TABLE}",
                headers=h,
                params={
                    "select": "term_id,locale,translated_term",
                    "limit": "50000",
                },
                timeout=20,
            )
            if t_resp.ok:
                all_trans = t_resp.json()
                if isinstance(all_trans, list):
                    for tr in all_trans:
                        tid = tr.get("term_id", "")
                        loc = tr.get("locale", "")
                        if tid and loc:
                            trans_map.setdefault(tid, {})[loc] = tr.get("translated_term", "")
        except Exception:
            pass

    for term in terms:
        tid = term.get("id", "")
        term["translations"] = trans_map.get(tid, {})

    return terms


def match_glossary_terms(text: str, glossary_terms: List[Dict]) -> List[Dict]:
    """
    Scan text for glossary source terms.
    Returns list of matched terms with positions.
    """
    matches = []
    if not text or not glossary_terms:
        return matches

    for term in glossary_terms:
        source = (term.get("source_term") or "").strip()
        if not source:
            continue
        pattern = re.compile(re.escape(source), re.IGNORECASE)
        found = pattern.findall(text)
        if found:
            matches.append({
                "term_id": term.get("id", ""),
                "glossary_id": term.get("glossary_id", ""),
                "source_term": source,
                "translations": term.get("translations", {}),
                "match_count": len(found),
            })

    return matches


def build_glossary_prompt(matched_terms: List[Dict], target_locale: str) -> str:
    """
    Build a glossary constraint section for the translation prompt.
    This is injected into the system prompt to enforce glossary terms.
    """
    if not matched_terms:
        return ""

    lines = ["GLOSSARY RULES (MUST follow these translations exactly):"]
    for m in matched_terms:
        source = m.get("source_term", "")
        translations = m.get("translations", {})
        target = translations.get(target_locale, "")
        if target:
            lines.append(f'  "{source}" -> "{target}"')
        else:
            lines.append(f'  "{source}" -> keep as "{source}" (do not translate)')

    lines.append("")
    lines.append("You MUST use the exact glossary translations above. Do not deviate.")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# XLSX Import / Export
# ---------------------------------------------------------------------------

def export_glossary_xlsx(glossary_id: str) -> bytes:
    """
    Export a glossary to XLSX format.
    Returns raw bytes of the XLSX file.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")

    glossary = get_glossary(glossary_id)
    if not glossary:
        raise ValueError("Glossary not found")

    target_locales = glossary.get("target_locales", [])
    if isinstance(target_locales, str):
        import json
        target_locales = json.loads(target_locales)

    # Get all terms with translations
    terms_data = list_terms(glossary_id, page=1, page_size=50000)
    terms = terms_data.get("terms", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"

    # Header row
    headers = ["Source Term", "Part of Speech", "Description"]
    for loc in target_locales:
        lang_name = TARGET_LANGUAGES.get(loc, loc)
        headers.append(f"{lang_name} ({loc})")
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for term in terms:
        row = [
            term.get("source_term", ""),
            term.get("part_of_speech", ""),
            term.get("description", ""),
        ]
        translations = term.get("translations", {})
        for loc in target_locales:
            row.append(translations.get(loc, ""))
        ws.append(row)

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value or ""))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def import_glossary_xlsx(glossary_id: str, file_bytes: bytes) -> Dict:
    """
    Import terms from XLSX into a glossary.
    Supports creating new terms and updating existing ones.
    Returns: { created, updated, errors }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for XLSX import. Install with: pip install openpyxl")

    glossary = get_glossary(glossary_id)
    if not glossary:
        raise ValueError("Glossary not found")

    target_locales = glossary.get("target_locales", [])
    if isinstance(target_locales, str):
        import json
        target_locales = json.loads(target_locales)

    buf = BytesIO(file_bytes)
    try:
        wb = load_workbook(buf, read_only=False)
    except Exception as e:
        return {"created": 0, "updated": 0, "errors": [f"Failed to open XLSX file: {str(e)}"]}
    
    ws = wb.active
    if not ws:
        return {"created": 0, "updated": 0, "errors": ["Workbook has no active sheet."]}

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"created": 0, "updated": 0, "errors": [f"File has no data rows. Found {len(rows)} row(s)."]}

    header = [str(c or "").strip() for c in rows[0]]

    # Find column indices
    source_idx = None
    pos_idx = None
    desc_idx = None
    locale_cols: Dict[str, int] = {}

    for i, col_name in enumerate(header):
        col_lower = col_name.lower().strip()
        # More flexible matching for source term column
        if col_lower in ("source term", "source", "term", "source_term", "sourceterm"):
            source_idx = i
        elif col_lower in ("part of speech", "pos", "part_of_speech", "partofspeech"):
            pos_idx = i
        elif col_lower in ("description", "desc", "note", "description", "descr"):
            desc_idx = i
        else:
            # Try to match locale - check for patterns like "Arabic (ar)", "ar", "Arabic", etc.
            for loc in target_locales:
                lang_name = TARGET_LANGUAGES.get(loc, loc)
                # Check multiple patterns
                if (loc.lower() in col_lower or 
                    lang_name.lower() in col_lower or
                    f"({loc.lower()})" in col_lower or
                    f"({loc.upper()})" in col_lower):
                    locale_cols[loc] = i
                    break

    if source_idx is None:
        return {
            "created": 0, 
            "updated": 0, 
            "errors": [f"Could not find 'Source Term' column in header. Found columns: {', '.join(header)}"]
        }

    # Get existing terms for duplicate check
    existing_terms = list_terms(glossary_id, page=1, page_size=50000)
    existing_by_source: Dict[str, Dict] = {}
    for t in existing_terms.get("terms", []):
        key = (t.get("source_term") or "").strip().lower()
        if key:
            existing_by_source[key] = t

    created = 0
    updated = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            # Convert row to list if it's a tuple
            if isinstance(row, tuple):
                row = list(row)
            
            # Ensure row has enough elements
            while len(row) <= max([source_idx, pos_idx or 0, desc_idx or 0] + [c for c in locale_cols.values()]):
                row.append(None)
            
            source_term = str(row[source_idx] or "").strip() if source_idx < len(row) and row[source_idx] is not None else ""
            if not source_term or source_term.lower() in ("", "none", "null"):
                continue

            pos = ""
            if pos_idx is not None and pos_idx < len(row) and row[pos_idx] is not None:
                pos = str(row[pos_idx] or "").strip()
            
            desc = ""
            if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None:
                desc = str(row[desc_idx] or "").strip()

            translations: Dict[str, str] = {}
            for loc, col_idx in locale_cols.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx] or "").strip()
                    if val and val.lower() not in ("", "none", "null"):
                        translations[loc] = val

            key = source_term.lower()
            if key in existing_by_source:
                # Update existing term
                existing = existing_by_source[key]
                term_id = existing.get("id", "")
                if term_id:
                    # Merge translations (don't wipe unspecified ones)
                    existing_trans = existing.get("translations", {})
                    merged = {**existing_trans, **translations}
                    update_term(term_id, {
                        "source_term": source_term,
                        "part_of_speech": pos or existing.get("part_of_speech", ""),
                        "description": desc or existing.get("description", ""),
                    }, merged)
                    updated += 1
            else:
                # Create new term
                create_term(
                    glossary_id=glossary_id,
                    source_term=source_term,
                    translations=translations,
                    part_of_speech=pos,
                    description=desc,
                )
                created += 1
                existing_by_source[key] = {"source_term": source_term}

        except Exception as e:
            import traceback
            error_msg = f"Row {row_idx}: {str(e)}"
            if len(errors) < 10:  # Limit error messages
                errors.append(error_msg)
            elif len(errors) == 10:
                errors.append(f"... and more errors (showing first 10)")

    return {"created": created, "updated": updated, "errors": errors}
